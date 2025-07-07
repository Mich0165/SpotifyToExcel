import os
import sys
import re
import spotipy
from spotipy.oauth2 import SpotifyOAuth
import pandas as pd
from dotenv import load_dotenv

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

load_dotenv()  # load environment variables from .env

CLIENT_ID = os.getenv("SPOTIPY_CLIENT_ID")
CLIENT_SECRET = os.getenv("SPOTIPY_CLIENT_SECRET")
REDIRECT_URI = os.getenv("SPOTIPY_REDIRECT_URI")
SCOPE = "playlist-read-private"

PLAYLIST_ID = None
if len(sys.argv) > 1:
    PLAYLIST_ID = sys.argv[1]
else:
    PLAYLIST_ID = os.getenv("SPOTIPY_PLAYLIST_ID")

if not PLAYLIST_ID:
    print("Error: Please provide a playlist ID as argument or set SPOTIPY_PLAYLIST_ID environment variable.")
    sys.exit(1)

sp_oauth = SpotifyOAuth(CLIENT_ID, CLIENT_SECRET, REDIRECT_URI, scope=SCOPE)
sp = spotipy.Spotify(auth_manager=sp_oauth)


def ms_to_time(ms):
    seconds = ms // 1000
    minutes = seconds // 60
    seconds = seconds % 60
    return f"{minutes}:{seconds:02}"


def clean_filename(s):
    return re.sub(r'[\\/*?:"<>|]', '', s)


def main():
    print(f"Fetching playlist {PLAYLIST_ID} info...")
    playlist = sp.playlist(PLAYLIST_ID)

    playlist_name = playlist['name']
    playlist_owner = playlist['owner']['display_name']
    playlist_url = playlist['external_urls']['spotify']

    print(f"Playlist: {playlist_name} by {playlist_owner}")

    results = sp.playlist_items(PLAYLIST_ID, additional_types=['track'])
    tracks = results['items']

    while results['next']:
        results = sp.next(results)
        tracks.extend(results['items'])

    data = []
    total_duration_ms = 0

    for item in tracks:
        track = item['track']
        if not track:
            continue
        track_name = track['name']
        artists = ", ".join([artist['name'] for artist in track['artists']])
        album = track['album']['name']
        duration_ms = track['duration_ms']
        total_duration_ms += duration_ms
        length = ms_to_time(duration_ms)
        track_url = track['external_urls']['spotify']

        data.append({
            "Track Name": track_name,
            "Artists": artists,
            "Album": album,
            "Length": length,
            "Track URL": track_url
        })

    df = pd.DataFrame(data)

    # Summary row
    total_minutes = total_duration_ms // 60000
    hours = total_minutes // 60
    minutes = total_minutes % 60
    summary = {
        "Track Name": f"Playlist: {playlist_name}",
        "Artists": f"Owner: {playlist_owner}",
        "Album": f"URL: {playlist_url}",
        "Length": f"Total Duration: {hours}h {minutes}m"
    }
    df = pd.concat([df, pd.DataFrame([summary])], ignore_index=True)

    # Create safe filename
    clean_name = clean_filename(playlist_name)
    clean_owner = clean_filename(playlist_owner)
    output_file = f"Playlist {clean_name} by {clean_owner}.xlsx"

    df.to_excel(output_file, index=False)

    # Excel formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    # Autofilter
    ws.auto_filter.ref = ws.dimensions

    # Set column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Add hyperlinks for track URLs
    url_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Track URL":
            url_col_idx = idx
            break

    if url_col_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=url_col_idx)
            if cell.value and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.value = "Link"
                cell.style = "Hyperlink"

    wb.save(output_file)
    print(f"Playlist data exported to {output_file}")


if __name__ == "__main__":
    main()
